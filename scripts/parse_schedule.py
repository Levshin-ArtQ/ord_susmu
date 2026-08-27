#!/usr/bin/env python3
"""Parse the ordinators' year schedule xlsx into compact JSON/JS seed."""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MONTHS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}

PRACTICE_FILL = "B6D7A8"
RED_FILLS = {"FF0000", "CC0000"}
EMPTY_FILLS = {None, "", "00000000", "000000", "FFFFFF", "FFFFFFFF"}

LETTER_RUN_PRACTICE = re.compile(r"^практика$", re.I)
LETTER_RUN_PA = re.compile(r"^па$", re.I)
PLUS_GROUPS = re.compile(r"\s*\+\s*(\d+(?:-\d+)?(?:\s*\+\s*\d+(?:-\d+)?)*)\s*$")
SINGLE_LETTER = re.compile(r"^[A-Za-zА-Яа-яЁё]$")


def rgb_of(cell) -> str | None:
    fill = cell.fill
    if not fill or fill.fgColor is None:
        return None
    fg = fill.fgColor
    if fg.type == "rgb" and fg.rgb:
        h = str(fg.rgb).upper()
        if h.startswith("FF") and len(h) == 8:
            h = h[2:]
        if h in ("00000000", "000000"):
            return None
        if len(h) == 8:
            h = h[2:]
        if h == "FFFFFF":
            return None
        return h
    if fg.type == "theme":
        # theme 0 (dk1) used as empty in this sheet
        if fg.theme in (0, 1) and not fg.tint:
            return None
        return None
    return None


def cell_text(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def normalize_title(raw: str) -> str:
    t = re.sub(r"\s+", " ", (raw or "").strip())
    t = t.replace("Ё", "Е").replace("ё", "е")
    return t


def base_title(raw: str) -> str:
    t = normalize_title(raw)
    t = PLUS_GROUPS.sub("", t).strip(" ,+")
    t = re.sub(r"\s+", " ", t)
    return t


def plus_groups(raw: str) -> list[str]:
    m = PLUS_GROUPS.search(normalize_title(raw))
    if not m:
        return []
    return re.findall(r"\d+(?:-\d+)?", m.group(1))


def expand_merges(ws, min_row, max_row, min_col, max_col):
    """Map every cell to the top-left master of its merge, if any."""
    master = {}
    for rng in ws.merged_cells.ranges:
        if rng.max_row < min_row or rng.min_row > max_row:
            continue
        if rng.max_col < min_col or rng.min_col > max_col:
            continue
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                master[(r, c)] = (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
    return master


def build_days(ws, first_col: int, last_col: int):
    month_header_at = {}
    for col in range(1, last_col + 1):
        v = cell_text(ws.cell(1, col).value).lower()
        if v in MONTHS:
            month_header_at[col] = MONTHS[v]

    days = []
    year = 2026
    month = 9
    last_day_num = 0
    seen_header_months = {9}

    for col in range(first_col, last_col + 1):
        raw = ws.cell(2, col).value
        if raw is None:
            continue
        if isinstance(raw, str) and ("-" in raw or "–" in raw or "каник" in raw.lower()):
            text = raw.strip()
            # Vacation range after the academic year. Sheet says 2026; context is 2027.
            m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})\s*[-–]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
            start = end = None
            if m:
                d1, mo1, y1, d2, mo2, y2 = map(int, m.groups())
                # If this column sits after July 2027, prefer 2027.
                if days and days[-1]["date"] >= "2027-07-01" and y1 == 2026:
                    y1, y2 = 2027, 2027
                start = date(y1, mo1, d1).isoformat()
                end = date(y2, mo2, d2).isoformat()
            days.append(
                {
                    "col": col,
                    "date": start or "2027-07-12",
                    "end": end or "2027-08-31",
                    "kind": "range",
                    "label": text,
                    "w": None,
                }
            )
            continue

        try:
            day_num = int(raw)
        except (TypeError, ValueError):
            continue

        header_month = month_header_at.get(col)
        if last_day_num and day_num < last_day_num:
            month += 1
            if month > 12:
                month = 1
                year += 1
        if header_month:
            # Header may appear 1 column after the real month start.
            if header_month != month:
                if not (last_day_num and day_num < last_day_num) and header_month == month + 1 or (
                    month == 12 and header_month == 1
                ):
                    month = header_month
                    if header_month == 1 and 9 in seen_header_months:
                        year = 2027
                elif header_month != month:
                    # Trust the day-rollover; keep month, remember header.
                    pass
            seen_header_months.add(header_month)

        # Clamp invalid dates (should not happen with rollover)
        try:
            iso = date(year, month, day_num)
        except ValueError:
            month = header_month or month
            iso = date(year, month, day_num)

        days.append(
            {
                "col": col,
                "date": iso.isoformat(),
                "end": None,
                "kind": "day",
                "label": None,
                "w": iso.weekday(),  # Mon=0
            }
        )
        last_day_num = day_num

    return days


def classify_raw(text: str, fill: str | None) -> tuple[str, str]:
    """Return (kind, title) from a single cell before letter-run analysis."""
    t = normalize_title(text)
    low = t.lower()
    if "каник" in low:
        return "vacation", "Каникулы"
    if low == "практика":
        return "practice", "Практика"
    if fill and fill.upper() in RED_FILLS:
        if t in ("П", "А", "ПА", "ПП"):
            return "attestation", t
        if not t:
            return "off", "Неучебный день"
        return "off", t
    if fill and fill.upper() == PRACTICE_FILL:
        if not t or SINGLE_LETTER.match(t):
            return "practice", "Практика"
        return "course", t
    if t and not SINGLE_LETTER.match(t) and t not in ("П", "А", "ПА"):
        return "course", t
    if t and SINGLE_LETTER.match(t):
        return "letter", t
    return "empty", ""


def spread_course_fills(cells: list[dict]):
    """Colored-but-empty cells continue the previous course of the same fill.

    The sheet often highlights a cycle without merging the cells; only the first
    cell holds the discipline name.
    """
    last = None  # (fill, title, plus, vshare)
    for c in cells:
        if c["kind"] == "course" and c.get("fill"):
            last = (c["fill"], c["title"], list(c.get("plus") or []), list(c.get("vshare") or []))
            continue
        if c["kind"] in ("empty", "letter") and last and c.get("fill") == last[0]:
            c["kind"] = "course"
            c["title"] = last[1]
            c["raw"] = last[1]
            c["plus"] = sorted(set((c.get("plus") or []) + last[2]))
            c["vshare"] = sorted(set((c.get("vshare") or []) + last[3]))
            continue
        if c["kind"] not in ("empty",):
            last = None


def apply_letter_runs(cells: list[dict]):
    i = 0
    n = len(cells)
    while i < n:
        if cells[i]["kind"] not in ("letter", "attestation") and not (
            cells[i]["kind"] == "off" and len(cells[i]["title"]) == 1
        ):
            i += 1
            continue
        j = i
        letters = []
        idxs = []
        while j < n:
            k = cells[j]["kind"]
            t = cells[j]["title"]
            if k in ("letter", "attestation") or (k == "off" and len(t) == 1):
                letters.append(t)
                idxs.append(j)
                j += 1
            elif k == "empty" and j + 1 < n and cells[j + 1]["kind"] in ("letter", "attestation"):
                # gap inside a spelled word — rare, stop
                break
            else:
                break
        joined = "".join(letters).lower().replace("ё", "е")
        if LETTER_RUN_PRACTICE.match(joined):
            for ix in idxs:
                cells[ix]["kind"] = "practice"
                cells[ix]["title"] = "Практика"
                cells[ix]["raw"] = "Практика"
        elif LETTER_RUN_PA.match(joined) or joined in ("па", "пп"):
            for ix in idxs:
                cells[ix]["kind"] = "attestation"
                cells[ix]["title"] = "Промежуточная аттестация"
                cells[ix]["raw"] = "ПА"
        i = j if j > i else i + 1


def coalesce_practice(cells: list[dict]):
    """Empty cells with practice fill, or empty cells inside a practice run, become practice."""
    n = len(cells)
    practice_idx = [i for i, c in enumerate(cells) if c["kind"] == "practice"]
    if not practice_idx:
        # still convert practice-colored empties
        for c in cells:
            if c.get("fill") == PRACTICE_FILL and c["kind"] in ("empty", "letter"):
                c["kind"] = "practice"
                c["title"] = "Практика"
                c["raw"] = "Практика"
        return

    start, end = min(practice_idx), max(practice_idx)
    # extend backwards/forwards over practice-colored empties
    while start > 0 and cells[start - 1].get("fill") == PRACTICE_FILL and cells[start - 1]["kind"] in (
        "empty",
        "letter",
        "practice",
    ):
        start -= 1
    while end + 1 < n and cells[end + 1].get("fill") == PRACTICE_FILL and cells[end + 1]["kind"] in (
        "empty",
        "letter",
        "practice",
    ):
        end += 1
    for i in range(start, end + 1):
        if cells[i]["kind"] in ("empty", "letter", "practice"):
            cells[i]["kind"] = "practice"
            cells[i]["title"] = "Практика"
            cells[i]["raw"] = "Практика"


def to_blocks(group_id: str, speciality: str, cells: list[dict]) -> list[dict]:
    blocks = []
    i = 0
    n = len(cells)
    while i < n:
        c = cells[i]
        kind = c["kind"]
        if kind == "empty":
            kind = "specialty"
            title = speciality
            color = None
        else:
            title = c["title"] or speciality
            color = c.get("fill")
            if kind == "letter":
                kind = "specialty"
                title = speciality
                color = None

        j = i
        while j + 1 < n:
            nxt = cells[j + 1]
            nkind = nxt["kind"]
            if nkind == "empty" or nkind == "letter":
                nkind2 = "specialty"
                ntitle = speciality
                ncolor = None
            else:
                nkind2 = nkind
                ntitle = nxt["title"] or speciality
                ncolor = nxt.get("fill")
            same = nkind2 == kind and base_title(ntitle) == base_title(title)
            if kind == "course":
                same = same and (ncolor == color or not ncolor or not color)
            if not same:
                break
            j += 1

        shared = []
        for k in range(i, j + 1):
            shared.extend(cells[k].get("plus") or [])
            shared.extend(cells[k].get("vshare") or [])
        shared = sorted({s for s in shared if s and s != group_id})

        start = cells[i]["date"]
        end = cells[j].get("end") or cells[j]["date"]
        work_days = sum(1 for k in range(i, j + 1) if not cells[k].get("range"))
        if work_days == 0 and start and end:
            try:
                d0 = date.fromisoformat(start)
                d1 = date.fromisoformat(end)
                work_days = (d1 - d0).days + 1
            except ValueError:
                work_days = 0

        blocks.append(
            {
                "id": f"{group_id}:{start}:{kind}:{base_title(title)[:40]}",
                "kind": kind,
                "title": title if kind != "specialty" else speciality,
                "base": base_title(title if kind != "specialty" else speciality),
                "color": color,
                "start": start,
                "end": end,
                "dayCount": work_days,
                "sharedHint": shared,
                "raw": c.get("raw") or title,
            }
        )
        i = j + 1
    return blocks


def parse(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    groups = []
    for row in range(3, 200):
        spec = cell_text(ws.cell(row, 1).value)
        grp = ws.cell(row, 2).value
        if not spec and grp is None:
            if groups and row > 10:
                break
            continue
        if grp is None:
            continue
        gid = cell_text(grp)
        groups.append({"id": gid, "speciality": normalize_title(spec), "row": row})

    last_col = 2
    for col in range(3, ws.max_column + 1):
        if ws.cell(2, col).value is not None:
            last_col = col
    days_meta = build_days(ws, 3, last_col)

    master = expand_merges(ws, 3, groups[-1]["row"], 3, last_col)

    # vertical share: other group ids in the same merge
    row_to_gid = {g["row"]: g["id"] for g in groups}

    parsed_groups = []
    for g in groups:
        row = g["row"]
        cells = []
        for d in days_meta:
            col = d["col"]
            key = (row, col)
            mrow, mcol = row, col
            vshare = []
            if key in master:
                r0, c0, r1, c1 = master[key]
                mrow, mcol = r0, c0
                for rr in range(r0, r1 + 1):
                    gid = row_to_gid.get(rr)
                    if gid and gid != g["id"]:
                        vshare.append(gid)
            cell = ws.cell(mrow, mcol)
            text = cell_text(cell.value)
            fill = rgb_of(cell) or rgb_of(ws.cell(row, col))
            kind, title = classify_raw(text, fill)
            cells.append(
                {
                    "date": d["date"],
                    "end": d.get("end"),
                    "range": d["kind"] == "range",
                    "kind": kind if d["kind"] != "range" or kind != "empty" else "vacation",
                    "title": "Каникулы" if d["kind"] == "range" and kind in ("empty", "vacation") else title,
                    "raw": text,
                    "fill": fill,
                    "plus": plus_groups(text),
                    "vshare": vshare,
                }
            )
        spread_course_fills(cells)
        apply_letter_runs(cells)
        coalesce_practice(cells)
        # vacation column
        for c in cells:
            if c.get("range"):
                c["kind"] = "vacation"
                c["title"] = "Каникулы"
        blocks = to_blocks(g["id"], g["speciality"], cells)
        parsed_groups.append(
            {
                "id": g["id"],
                "speciality": g["speciality"],
                "blocks": blocks,
            }
        )

    days = []
    for d in days_meta:
        if d["kind"] == "range":
            days.append(
                {
                    "date": d["date"],
                    "end": d["end"],
                    "w": d["w"],
                    "kind": "range",
                    "label": d["label"],
                }
            )
        else:
            days.append({"date": d["date"], "w": d["w"], "kind": "day"})

    return {
        "version": 1,
        "academicYear": "2026-2027",
        "yearLabel": "1 год",
        "title": "Расписание ординатуры 2026–2027",
        "sourceName": path.name,
        "groups": parsed_groups,
        "days": days,
    }


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "/Users/artemlevsin/Downloads/Расписание 2026-2027 1 год .xlsx")
    out_dir = Path(sys.argv[2] if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent / "data")
    data = parse(src)
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "schedule.json"
    js_path = out_dir / "seed.js"
    json_path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    js_path.write_text(
        "window.SEED=" + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )
    print("groups", len(data["groups"]))
    print("days", len(data["days"]))
    print("wrote", json_path, "bytes", json_path.stat().st_size)
    print("wrote", js_path, "bytes", js_path.stat().st_size)

    g = next(x for x in data["groups"] if x["id"] == "101-1")
    print("\n=== 101-1 blocks ===")
    for b in g["blocks"]:
        print(f"  {b['start']}..{b['end']} [{b['kind']:12}] n={b['dayCount']:3} {b['title'][:60]!r} #{b['color']} share={b['sharedHint']}")

    # date sanity
    seq = [d["date"] for d in data["days"] if d["kind"] == "day"]
    print("\nfirst", seq[0], "last", seq[-1], "count", len(seq))
    dups = [d for d in seq if seq.count(d) > 1]
    print("duplicate dates", sorted(set(dups))[:10], "n", len(set(dups)))
    # check monotonic
    bad = [(a, b) for a, b in zip(seq, seq[1:]) if a >= b]
    print("non-increasing", bad[:8], "n", len(bad))


if __name__ == "__main__":
    main()
